"""
情熱AIプロジェクト - Excel出力モジュール
openpyxlを使用して圃場一覧レポートを生成する
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# 色定義
HEADER_BG = "1F497D"      # 濃い青
HEADER_FG = "FFFFFF"      # 白
ROW_ODD_BG = "FFFFFF"     # 白
ROW_EVEN_BG = "DCE6F1"    # 薄い青

COLUMNS = [
    ("圃場ID",      "id",           10),
    ("圃場名",      "name",         20),
    ("住所",        "address",      30),
    ("解析日",      "analyzed_at",  18),
    ("土壌タイプ",  "soil_type",    16),
    ("荒廃度",      "abandonment_level", 10),
    ("栽培適性TOP3","top3_crops",   35),
]


def _format_top3(crops: list[dict[str, Any]] | None) -> str:
    """crop_suitabilityリストからTOP3文字列を生成する。"""
    if not crops:
        return "-"
    sorted_crops = sorted(crops, key=lambda c: c.get("rank", 999))[:3]
    parts = []
    for crop in sorted_crops:
        name = crop.get("crop_name", "不明")
        score = crop.get("score", 0)
        parts.append(f"{name}({score}点)")
    return " / ".join(parts)


def _thin_border() -> Border:
    side = Side(style="thin", color="AAAAAA")
    return Border(left=side, right=side, top=side, bottom=side)


def generate_report(fields_data: list[dict[str, Any]]) -> bytes:
    """
    圃場データリストからExcelレポートを生成し、bytesで返す。

    Parameters
    ----------
    fields_data : list[dict]
        各要素は圃場情報を含む辞書。以下のキーを期待:
          - id, name, address
          - analyzed_at (str | None)
          - soil_type (str | None)
          - abandonment_level (int | None)
          - crop_suitability (list[dict] | None)
            各dict: crop_name, score, rank

    Returns
    -------
    bytes
        Excelファイルのバイナリ
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "圃場一覧"

    # --- ヘッダー行 ---
    header_font = Font(name="Meiryo", bold=True, color=HEADER_FG, size=11)
    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_BG)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (label, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # --- データ行 ---
    border = _thin_border()
    data_align_center = Alignment(horizontal="center", vertical="center")
    data_align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for row_idx, field in enumerate(fields_data, start=2):
        is_even = (row_idx % 2 == 0)
        fill = PatternFill(fill_type="solid", fgColor=ROW_EVEN_BG if is_even else ROW_ODD_BG)
        font = Font(name="Meiryo", size=10)

        values = [
            field.get("id", ""),
            field.get("name", ""),
            field.get("address", ""),
            field.get("analyzed_at", "-") or "-",
            field.get("soil_type", "-") or "-",
            field.get("abandonment_level", "-") if field.get("abandonment_level") is not None else "-",
            _format_top3(field.get("crop_suitability")),
        ]

        for col_idx, (_, _, _), value in zip(range(1, len(COLUMNS) + 1), COLUMNS, values):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = font
            cell.fill = fill
            cell.border = border
            # 住所・栽培適性は左揃え、それ以外は中央
            if col_idx in (3, 7):
                cell.alignment = data_align_left
            else:
                cell.alignment = data_align_center

        ws.row_dimensions[row_idx].height = 18

    # --- シート外観調整 ---
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = True

    # --- bytesとして返す ---
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
